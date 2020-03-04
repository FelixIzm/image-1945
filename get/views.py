from django.shortcuts import render
from django.http import HttpResponse
from .forms import UserForm
import os, tempfile, requests, time
#import get_image_google

import requests, json, re
import hashlib
import lxml.html as html
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload,MediaFileUpload
from googleapiclient.discovery import build
from string import Template
from pathlib import Path
import tempfile
import pprint
from openpyxl import Workbook
from datetime import datetime
import pytz


image_id =''
##########################################
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

cols = ['ID scan','ID','Фамилия','Имя','Отчество','Дата рождения/Возраст','Место рождения','Дата и место призыва','Последнее место службы','Воинское звание','Судьба','Дата смерти','Первичное место захоронения']

SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'obd.json'
credentials = service_account.Credentials.from_service_account_file(os.path.join(BASE_DIR, SERVICE_ACCOUNT_FILE), scopes=SCOPES)
service = build('drive', 'v3', credentials=credentials)
name_root_folder = 'Folder'

pp = pprint.PrettyPrinter(indent=4)

root_results = service.files().list(pageSize=10,fields="nextPageToken, files(id, name, mimeType,webViewLink)",q=Template("name contains '$name_root_folder'").safe_substitute(name_root_folder=name_root_folder)).execute()
id_root_folder = root_results['files'][0]['id']

######################################
def parse_file (name_file):
    dict_ = {}
    f = open(name_file, 'r')
    s = f.read()
    dict_={}
    list_ = s.splitlines()
    for item in list_:
        items = item.split(":")
        dict_[items[0]] = items[1].lstrip()
    return dict_
#####################################
def make_str_cookie(cookies):
    str_cook = ''
    for key, value in cookies.items():
        str_cook += '{0}={1};'.format(key,value)
    return str_cook
#####################################
def getStringHash(id):
    h = hashlib.md5(str(id).encode()+b'db76xdlrtxcxcghn7yusxjcdxsbtq1hnicnaspohh5tzbtgqjixzc5nmhybeh')
    p = h.hexdigest()
    return str(p)
#####################################
def get_info(id_scan,id,cookies):
    cookies['showimage']='0'
    info_url = 'https://obd-memorial.ru/html/info.htm?id='+str(id)
    res3 = requests.get(info_url,cookies=cookies)
    doc = html.fromstring(res3.text)
    divs = {}
    for div in doc.find_class('card_parameter'):
        divs[div.getchildren()[0].text_content()] = div.getchildren()[1].text_content()
        #print ('%s: %s' % (div.getchildren()[0].text_content(), div.getchildren()[1].text_content()))
    list_col = []
    for col in cols:
        if(col in divs.keys()):
            list_col.append(divs[col])
        else:
            list_col.append('')
    list_col[1] = id
    list_col[0] = id_scan
    return list_col

#####################################
#            MAIN!!!!!!!!!!         #
#####################################

def main(image_id,image,excel):
    if(image_id==None):
        return 'webViewLink'
    info_url = 'https://obd-memorial.ru/html/info.htm?id={}'.format(image_id)
    img_info = 'https://obd-memorial.ru/html/getimageinfo?id={}'.format(image_id)
    print(info_url)
    res1 = requests.get(info_url,allow_redirects = True)
    dirpath = tempfile.mkdtemp()
    print('dirpath = '+dirpath)
    # создаем каталог сразу - один раз
    #name_folder_save = str(image_id)+"_"+os.path.basename(tempfile.mktemp ())
    d = datetime.now()
    #.strftime('%Y-%m-%d:%H_%M_%S')
    #print(d.tzinfo) # Return time zone info
    d = pytz.timezone('Europe/Paris').localize(d)
    print(d.strftime('%Y-%m-%d %H:%M:%S'))
    name_folder_save = str(image_id)+' '+d.strftime('%Y-%m-%d %H:%M:%S')
    print('name_folder_save = '+name_folder_save)
    #create catalog
    file_metadata = {
        'name': name_folder_save,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [id_root_folder]
    }
    result = service.files().create(body=file_metadata, fields='id,webViewLink').execute()
    # id каталога для сохранения
    id_folder_save = result['id']
    # ссылка на каталог
    web_link = result['webViewLink']



    if(res1.status_code==307):
        print(res1.status_code)
        print('*****************')
        if(not '3fbe47cd30daea60fc16041479413da2' in res1.cookies):
            # Удаляем каталог за ненадобностью
            result = service.files().delete(fileId=id_folder_save).execute()
            print('*****************************************')
            print(' delete catalog = '+name_folder_save)
            pp.pprint(result)
            return 'Запись сводного документа не найдена'
        cookies = {}
        cookies['3fbe47cd30daea60fc16041479413da2']=res1.cookies['3fbe47cd30daea60fc16041479413da2']
        cookies['JSESSIONID']=res1.cookies['JSESSIONID']
        #############################
        #   load list id's images   #
        #############################
        response = requests.get(img_info,cookies=cookies)
        response_dict = json.loads(response.text)
        print('response_dict = '+str(len(response_dict)))
        #############################
        i=0
        if(excel):
            row_num = 1
            workbook = Workbook()
            # Get active worksheet/tab
            worksheet = workbook.active
            worksheet.title = 'Person'
            columns = cols
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
        # идем по списку id сканов
        # сохраняем имена файлов [id].jpg в list, потом его вернем для выгрузки Google Drive
        list_file = []
        for item in response_dict:
            i+=1
            #print(i, item['id'])
            if(excel):
                for id in item['mapData'].keys():
                    row_num += 1
                    row = get_info(item['id'],id,cookies)
                    #print('\t',id)
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value
            if(image):
                img_url="https://obd-memorial.ru/html/images3?id="+str(item['id'])+"&id1="+(getStringHash(item['id']))+"&path="+item['img']
                headers_302 = parse_file(BASE_DIR+'/header_302.txt')
                headers_302['Cookie'] = make_str_cookie(cookies)
                headers_302['Referer'] = info_url
                req302 = requests.get(img_url,headers=headers_302,cookies=cookies, allow_redirects = False)
                if(req302.status_code==302):
                    params = {}
                    params['id'] = str(item['id'])
                    params['id1'] = getStringHash(item['id'])
                    params['path'] = item['img']
                    headers_img = parse_file(BASE_DIR+'/header_img.txt')
                    headers_img['Referer'] = info_url
                    #####################
                    req_img = requests.get("https://cdn.obd-memorial.ru/html/images3",headers=headers_img,params=params,cookies=cookies,stream = True,allow_redirects = False )
                    #####################
                    if(req_img.status_code==200):
                        location = os.path.abspath(dirpath+"/"+str(item['id'])+'.jpg')
                        f = open(location, 'wb')
                        f.write(req_img.content)
                        f.close()
                        list_file.append(dirpath+"/"+str(item['id'])+'.jpg')

                        name = str(item['id'])+'.jpg'
                        print(name)
                        file_metadata = {'name': name,'parents': [id_folder_save]}
                        media = MediaFileUpload(dirpath+"/"+str(item['id'])+'.jpg', resumable=True,chunksize=-1, mimetype = 'image/jpg')
                        r = service.files().insert(body=file_metadata, media_body=media, fields='id').execute()


                        if(excel):
                            workbook.save(filename = (str(image_id)+'_book.xlsx'))


        # загружаем файлы на GoogleDrive
        '''
        for _file in list_file:
            name = os.path.basename(_file)
            print(_file)
            file_metadata = {'name': name,'parents': [id_folder_save]}
            media = MediaFileUpload(_file, resumable=True,chunksize=-1, mimetype = 'image/jpg')
            r = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        # Определяем - записались ли файлы в каталог
        '''
        result = service.files().list(pageSize=1000,fields="nextPageToken, files(id, name, mimeType,webViewLink)",q=Template("name contains '$name_folder_save'").safe_substitute(name_folder_save=name_folder_save)).execute()
        if(result['files']):
            return web_link
        else:
            return 'records not found'

##########################################
def index(request):
    print(" __name__ = "+str(__name__))
    #return HttpResponse("Hello, world. You're at the polls index.")
    image_id=0
    userform = UserForm({'image_id':image_id})
    image_id = request.POST.get("image_id")

    #image_id = 85942988 
    # #51480906 Иванов 2 скана
    link = ''
    d = {'image':True, 'excel':False}
    link = main(image_id,**d)
    return render(request, "get/index.html", {"form": userform,"web_link": link})

# Create your views here.
